#!/usr/bin/env python3
"""Export all UI catalog products across tribes to a structured Excel file."""

import json
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "src" / "data"
OUTPUT = ROOT / "sapiens_benchmarking_products.xlsx"
TITLES_PATH = DATA_DIR / "product_titles.json"

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WRAP = Alignment(wrap_text=True, vertical="top")


def load_product_titles() -> dict[str, str]:
    if not TITLES_PATH.is_file():
        return {}
    data = json.loads(TITLES_PATH.read_text())
    raw = data.get("titles") if isinstance(data, dict) else data
    if not isinstance(raw, dict):
        return {}
    return {str(k): str(v).strip() for k, v in raw.items() if str(v).strip()}


def load_catalog_rows():
    catalog = json.loads((DATA_DIR / "catalog-index.json").read_text())
    tribes_dir = DATA_DIR / "tribes"
    titles = load_product_titles()
    rows = []

    for tribe_meta in catalog["tribes"]:
        tribe_id = tribe_meta["id"]
        tribe = json.loads((tribes_dir / f"{tribe_id}.json").read_text())
        tribe_high_count = tribe_meta.get("highPriorityCount")
        tribe_max_high_gap = tribe_meta.get("maxHighPriorityGap")

        for user in tribe.get("member_user_characteristics", []):
            user_id = user["user_id"]
            char_summary = user.get("characteristic_summary", "")
            products = tribe.get("members_grouped_by_user", {}).get(user_id, [])

            for idx, p in enumerate(products, start=1):
                product_desc = (p.get("product_description") or "").strip()
                review_key = p.get("review_key", "")
                domain = tribe.get("domain", tribe_meta.get("domain", ""))
                main_desc = (p.get("main_product_description") or "").strip()
                if domain == "video_games":
                    product_title = titles.get(review_key, "")
                    ui_display_name = product_title or product_desc
                else:
                    product_title = ""
                    ui_display_name = main_desc if main_desc and main_desc != product_desc else product_desc

                rows.append(
                    {
                        "domain": domain,
                        "tribe_id": tribe_id,
                        "tribe_name": tribe.get("tribe_name", tribe_meta["name"]),
                        "cluster": tribe_meta.get("cluster", ""),
                        "micro_id": tribe_meta.get("microId", ""),
                        "user_id": user_id,
                        "user_characteristic_summary": char_summary,
                        "product_rank_in_user": idx,
                        "catalog_priority_tier": p.get("catalog_priority_tier", ""),
                        "high_priority_count": tribe_high_count,
                        "max_high_priority_gap": tribe_max_high_gap,
                        "review_key": review_key,
                        "product_title": product_title,
                        "ui_product_name": ui_display_name,
                        "product_description": product_desc,
                        "main_product_description": main_desc if domain == "healthcare" else "",
                        "category": p.get("category", ""),
                        "major_subcategory": p.get("major_subcategory", ""),
                        "sentiment": p.get("sentiment", ""),
                        "rating": p.get("rating"),
                        "sapiens_baseline_gap": p.get("sapiens_baseline_gap"),
                        "overall_similarity_score": p.get("overall_similarity_score"),
                        "healthcare_benchmark": p.get("healthcare_benchmark", False),
                        "video_games_benchmark": p.get("video_games_benchmark", False),
                        "tribe_mean_gap": tribe_meta.get("meanSapiensBaselineGap"),
                        "tribe_mean_similarity": tribe_meta.get("meanSimilarityScore"),
                        "tribe_product_count": tribe_meta.get("reviewCount"),
                        "tribe_user_count": tribe_meta.get("userCount"),
                    }
                )

    return rows, catalog


def style_header_row(ws, row_num, col_count):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def auto_width(ws, min_width=10, max_width=60):
    for col_cells in ws.columns:
        length = max(len(str(c.value or "")) for c in col_cells)
        adjusted = min(max(length + 2, min_width), max_width)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = adjusted


def write_sheet(ws, headers, data_rows):
    ws.append(headers)
    style_header_row(ws, 1, len(headers))
    for row in data_rows:
        ws.append(row)
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = WRAP
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions


def build_workbook(rows, catalog):
    wb = Workbook()

    # --- Sheet 1: Tribe Summary ---
    ws_summary = wb.active
    ws_summary.title = "Tribe Summary"
    summary_headers = [
        "Domain",
        "Tribe ID",
        "Tribe Name",
        "Cluster",
        "Micro ID",
        "Users in UI",
        "Products in UI",
        "Mean Sapiens-Baseline Gap",
        "Mean Similarity Score",
        "Description",
    ]
    summary_rows = []
    for t in sorted(catalog["tribes"], key=lambda x: (x.get("domain", ""), x["name"])):
        summary_rows.append(
            [
                t.get("domain", ""),
                t["id"],
                t["name"],
                t.get("cluster", ""),
                t.get("microId", ""),
                t.get("userCount", 0),
                t.get("reviewCount", 0),
                t.get("meanSapiensBaselineGap"),
                t.get("meanSimilarityScore"),
                t.get("description", ""),
            ]
        )
    write_sheet(ws_summary, summary_headers, summary_rows)
    auto_width(ws_summary, max_width=50)
    ws_summary.column_dimensions["J"].width = 70

    # --- Sheet 2: All Products (main deliverable) ---
    ws_products = wb.create_sheet("All Products")
    product_headers = [
        "Domain",
        "Tribe Name",
        "Tribe ID",
        "User ID",
        "Product Rank (per user)",
        "Priority Tier",
        "Product Title",
        "UI Product Name",
        "Product Description",
        "Main Product Description",
        "Category",
        "Major Subcategory",
        "Review Key",
        "Sapiens Baseline Gap",
        "Overall Similarity Score",
        "Sentiment",
        "Rating",
        "Healthcare Benchmark",
        "Video Games Benchmark",
        "Tribe Mean Gap",
        "User Characteristic Summary",
    ]
    sorted_rows = sorted(
        rows,
        key=lambda r: (
            r["domain"] != "video_games",
            -(r.get("high_priority_count") or 0),
            -(r.get("max_high_priority_gap") or 0),
            {"high": 0, "medium": 1, "low": 2}.get(str(r.get("catalog_priority_tier") or "medium").lower(), 1),
            -(r["sapiens_baseline_gap"] or 0),
            r["tribe_name"],
            r["user_id"],
            r["review_key"],
        ),
    )
    by_tribe_rows = sorted(
        rows,
        key=lambda r: (
            r["domain"],
            r["tribe_name"],
            {"high": 0, "medium": 1, "low": 2}.get(str(r.get("catalog_priority_tier") or "medium").lower(), 1),
            -(r["sapiens_baseline_gap"] or 0),
            r["user_id"],
            r["review_key"],
        ),
    )
    product_data = [
        [
            r["domain"],
            r["tribe_name"],
            r["tribe_id"],
            r["user_id"],
            r["product_rank_in_user"],
            (r.get("catalog_priority_tier") or "").title(),
            r["product_title"],
            r["ui_product_name"],
            r["product_description"],
            r["main_product_description"] if r["main_product_description"] != r["product_description"] else "",
            r["category"],
            r["major_subcategory"],
            r["review_key"],
            r["sapiens_baseline_gap"],
            r["overall_similarity_score"],
            r["sentiment"],
            r["rating"],
            "Yes" if r["healthcare_benchmark"] else "",
            "Yes" if r["video_games_benchmark"] else "",
            r["tribe_mean_gap"],
            r["user_characteristic_summary"],
        ]
        for r in sorted_rows
    ]
    write_sheet(ws_products, product_headers, product_data)
    auto_width(ws_products, max_width=55)
    ws_products.column_dimensions["F"].width = 45
    ws_products.column_dimensions["G"].width = 55
    ws_products.column_dimensions["S"].width = 50

    # --- Sheet 3: Products by Tribe (pivot-style) ---
    ws_by_tribe = wb.create_sheet("Products by Tribe")
    by_tribe_headers = [
        "Tribe Name",
        "Domain",
        "Product #",
        "Product Title",
        "UI Product Name",
        "Category",
        "Sapiens Baseline Gap",
        "Overall Similarity Score",
    ]
    by_tribe_data = []
    current_tribe = None
    product_num = 0
    for r in by_tribe_rows:
        if r["tribe_name"] != current_tribe:
            current_tribe = r["tribe_name"]
            product_num = 0
        product_num += 1
        by_tribe_data.append(
            [
                r["tribe_name"],
                r["domain"],
                product_num,
                r["product_title"],
                r["ui_product_name"],
                r["category"],
                r["sapiens_baseline_gap"],
                r["overall_similarity_score"],
            ]
        )
    write_sheet(ws_by_tribe, by_tribe_headers, by_tribe_data)
    auto_width(ws_by_tribe, max_width=55)
    ws_by_tribe.column_dimensions["D"].width = 55

    # --- Sheet 4: Healthcare only ---
    ws_hc = wb.create_sheet("Healthcare Products")
    hc_rows = [r for r in sorted_rows if r["domain"] == "healthcare"]
    hc_data = [
        [
            r["tribe_name"],
            r["ui_product_name"],
            r["category"],
            r["major_subcategory"],
            r["sapiens_baseline_gap"],
            r["overall_similarity_score"],
            r["review_key"],
        ]
        for r in hc_rows
    ]
    write_sheet(
        ws_hc,
        ["Tribe Name", "UI Product Name", "Category", "Major Subcategory", "Sapiens Baseline Gap", "Similarity Score", "Review Key"],
        hc_data,
    )
    auto_width(ws_hc, max_width=55)
    ws_hc.column_dimensions["B"].width = 55

    # --- Sheet 5: Video Games only ---
    ws_vg = wb.create_sheet("Video Games Products")
    vg_rows = [r for r in sorted_rows if r["domain"] == "video_games"]
    vg_data = [
        [
            r["tribe_name"],
            r["product_title"],
            (r.get("catalog_priority_tier") or "").title(),
            r["category"],
            r["sapiens_baseline_gap"],
            r["overall_similarity_score"],
            r["review_key"],
        ]
        for r in vg_rows
    ]
    write_sheet(
        ws_vg,
        ["Tribe Name", "UI Product Name", "Priority Tier", "Category", "Sapiens Baseline Gap", "Similarity Score", "Review Key"],
        vg_data,
    )
    auto_width(ws_vg, max_width=55)
    ws_vg.column_dimensions["B"].width = 55

    # --- Sheet 6: Metadata ---
    ws_meta = wb.create_sheet("About")
    ws_meta["A1"] = "Sapiens Benchmarking — Product Catalog Export"
    ws_meta["A1"].font = Font(bold=True, size=14)
    ws_meta["A3"] = "Generated"
    ws_meta["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M")
    ws_meta["A4"] = "Total Tribes"
    ws_meta["B4"] = catalog["total"]
    ws_meta["A5"] = "Total Products (UI)"
    ws_meta["B5"] = len(rows)
    ws_meta["A6"] = "Healthcare Products"
    ws_meta["B6"] = len(hc_rows)
    ws_meta["A7"] = "Video Games Products"
    ws_meta["B7"] = len(vg_rows)
    ws_meta["A9"] = "Notes"
    ws_meta["A9"].font = Font(bold=True)
    notes = [
        "All products listed are shown in the benchmarking UI.",
        "Products are pre-filtered to cases where Sapiens outperforms the best baseline.",
        "All products must have Sapiens overall similarity >=65%.",
        "All products must beat the best baseline by >10 percentage points.",
        "UI Product Name = product_description field shown in the product picker (Step 2).",
        "Sapiens Baseline Gap = Sapiens overall similarity score minus best baseline overall similarity score.",
        "Main product sheets sort video games by tribe high-priority count, then High/Medium/Low tier, then Sapiens Baseline Gap.",
    ]
    for i, note in enumerate(notes, start=10):
        ws_meta[f"A{i}"] = note
        ws_meta[f"A{i}"].alignment = WRAP
    ws_meta.column_dimensions["A"].width = 90

    return wb


def main():
    rows, catalog = load_catalog_rows()
    wb = build_workbook(rows, catalog)
    wb.save(OUTPUT)
    print(f"Exported {len(rows)} products across {catalog['total']} tribes to:")
    print(OUTPUT)


if __name__ == "__main__":
    main()
