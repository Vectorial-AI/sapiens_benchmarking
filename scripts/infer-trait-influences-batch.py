#!/usr/bin/env python3
"""
Batch infer which user/tribe traits shaped SAPIENS predictions for UI catalog products.

Mirrors src/lib/inferred-traits-explanation.ts (gpt-4o-mini, evidence-grounded).

Prediction sources:
  - video_games tribes: blind_run_i2.json (cluster_0 micro_0, micro_9, micro_11)
  - healthcare tribes: healthcare_digital_technical_accuracy.json

Context sent to the model:
  - SAPIENS generated review (prediction only)
  - Tribe qualitative traits
  - User characteristic summary + category-specific characteristics

NOT sent: ground-truth review, user norms.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import time
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
WORKSPACE = ROOT.parent
DATA_DIR = ROOT / "src" / "data"
OUTPUTS = WORKSPACE / "outputs"
CLUSTERING = WORKSPACE / "Clustering"
USER_CHARS_PATH = WORKSPACE / "Data extraction" / "user_llm_characteristics.json"
CATEGORY_MAPPING_PATH = WORKSPACE / "Data extraction" / "category_mapping_to_7_main.json"
HEALTHCARE_ACCURACY_PATH = (
    CLUSTERING / "Prediction_Accuracy_Refined" / "healthcare_digital_technical_accuracy.json"
)

INFERRED_TRAITS_MODEL = "gpt-4o-mini"
HEALTH_MAIN = "Health & Personal Care"
VIDEO_GAMES_MAIN = "Video Games"
SOFTWARE_MAIN = "Software"

TRAIT_GROUP_LABELS = {
    "inherent_behavioral_traits": "Inherent behavioral traits",
    "latent_motivations": "Latent motivations",
    "validation_triggers": "Validation triggers",
    "friction_points": "Friction points",
    "implicit_goals": "Implicit goals",
}

VG_BLIND_RUN_PATHS = {
    ("cluster_0", "micro_0"): OUTPUTS / "amazon_sgo_health_care/cluster_0/micro_0/blind_run_i2.json",
    ("cluster_0", "micro_9"): OUTPUTS / "amazon_sgo_health_care/cluster_0/micro_9/blind_run_i2.json",
    ("cluster_0", "micro_11"): OUTPUTS / "amazon_sgo_health_care/cluster_0/micro_11/blind_run_i2.json",
}


@dataclass
class TraitCatalogEntry:
    trait: str
    source: str  # tribe | user
    trait_group: str | None = None


@dataclass
class CatalogProductRow:
    domain: str
    tribe_id: str
    tribe_name: str
    cluster: str
    micro_id: str
    user_id: str
    review_key: str
    product_description: str
    category: str
    sapiens_review: str
    user_characteristic_summary: str
    category_characteristics: dict[str, str]
    qualitative_summary: dict[str, Any]
    sapiens_baseline_gap: float | None = None
    overall_similarity_score: float | None = None
    summary: str = ""
    influences: list[dict[str, Any]] = field(default_factory=list)
    inference_source: str = ""


def load_json(path: Path) -> Any:
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def normalize_trait_key(text: str) -> str:
    return " ".join(str(text or "").split()).casefold()


def trait_text(item: Any) -> str:
    if isinstance(item, str):
        return item.strip()
    if isinstance(item, dict):
        return str(item.get("text") or "").strip()
    return ""


def non_empty_traits(items: list[Any]) -> list[str]:
    return [t for t in (trait_text(x) for x in items) if t]


def load_category_mapping() -> dict[str, str]:
    if not CATEGORY_MAPPING_PATH.exists():
        return {}
    return load_json(CATEGORY_MAPPING_PATH).get("category_to_main_mapping") or {}


def map_category_to_main(category: str, sub_to_main: dict[str, str]) -> str:
    return sub_to_main.get(category, category)


def load_user_category_characteristics() -> dict[str, dict[str, str]]:
    if not USER_CHARS_PATH.exists():
        return {}
    raw = load_json(USER_CHARS_PATH)
    out: dict[str, dict[str, str]] = {}
    for uid, user_data in raw.items():
        if not isinstance(user_data, dict):
            continue
        cat_chars = user_data.get("category_characteristics") or {}
        summaries: dict[str, str] = {}
        for main_cat, block in cat_chars.items():
            if isinstance(block, dict):
                summary = str(block.get("influencing_characteristics_summary") or "").strip()
                if summary:
                    summaries[str(main_cat)] = summary
        if summaries:
            out[str(uid)] = summaries
    return out


def domain_category_characteristics(
    user_id: str,
    user_cat_chars: dict[str, dict[str, str]],
    domain: str,
) -> dict[str, str]:
    if domain == "video_games":
        out: dict[str, str] = {}
        for main in (VIDEO_GAMES_MAIN, SOFTWARE_MAIN):
            text = (user_cat_chars.get(user_id) or {}).get(main, "").strip()
            if text:
                out[main] = text
        return out
    text = (user_cat_chars.get(user_id) or {}).get(HEALTH_MAIN, "").strip()
    return {HEALTH_MAIN: text} if text else {}


def format_user_characteristics(
    user_summary: str,
    category_characteristics: dict[str, str],
    category: str,
    sub_to_main: dict[str, str],
) -> str:
    lines: list[str] = []
    general = user_summary.strip()
    if general:
        lines.append(f"[General Characteristics] {general}")
    main = map_category_to_main(category, sub_to_main)
    cat_char = category_characteristics.get(main, "").strip()
    if cat_char:
        lines.append(f"[{main} Specific] {cat_char}")
    return "\n\n".join(lines) if lines else "(none)"


def extract_qualitative_traits(q: dict[str, Any]) -> dict[str, list[str]]:
    latent = q.get("latent_motivations") or {}
    latent_main = latent.get("main") if isinstance(latent, dict) else latent
    return {
        "inherent_behavioral_traits": non_empty_traits(q.get("inherent_behavioral_traits") or []),
        "latent_motivations": non_empty_traits(latent_main or []),
        "validation_triggers": non_empty_traits(q.get("validation_triggers") or []),
        "friction_points": non_empty_traits(q.get("friction_points") or []),
        "implicit_goals": non_empty_traits(q.get("implicit_goals") or []),
    }


def format_tribe_trait_catalog(traits: dict[str, list[str]]) -> str:
    sections: list[str] = []
    for key, label in TRAIT_GROUP_LABELS.items():
        items = traits.get(key) or []
        if items:
            numbered = "\n".join(f"  {i + 1}. {t}" for i, t in enumerate(items))
            sections.append(f"{label}:\n{numbered}")
    return "\n\n".join(sections)


def first_sentence(text: str, max_len: int = 160) -> str:
    trimmed = " ".join(str(text or "").split()).strip()
    if not trimmed:
        return ""
    match = re.match(r"^[^.!?]+[.!?]?", trimmed)
    sentence = (match.group(0) if match else trimmed).strip()
    if len(sentence) <= max_len:
        return sentence
    return sentence[: max_len - 1].rstrip() + "…"


def build_trait_catalog(
    traits: dict[str, list[str]],
    user_block: str,
    user_summary: str,
    category: str,
    category_characteristics: dict[str, str],
    sub_to_main: dict[str, str],
) -> tuple[str, str, list[TraitCatalogEntry]]:
    catalog: list[TraitCatalogEntry] = []
    for key, label in TRAIT_GROUP_LABELS.items():
        for trait in traits.get(key) or []:
            catalog.append(TraitCatalogEntry(trait=trait, source="tribe", trait_group=label))

    main = map_category_to_main(category, sub_to_main)
    category_trait = str(category_characteristics.get(main) or "").strip()
    if category_trait:
        catalog.append(
            TraitCatalogEntry(
                trait=first_sentence(category_trait, 180),
                source="user",
                trait_group="User traits",
            )
        )
    elif user_summary.strip():
        catalog.append(
            TraitCatalogEntry(
                trait=first_sentence(user_summary, 160),
                source="user",
                trait_group="User traits",
            )
        )
    return format_tribe_trait_catalog(traits), user_block, catalog


def resolve_catalog_entry(trait: str, catalog: list[TraitCatalogEntry]) -> TraitCatalogEntry | None:
    key = normalize_trait_key(trait)
    for entry in catalog:
        if normalize_trait_key(entry.trait) == key:
            return entry
    for entry in catalog:
        ek = normalize_trait_key(entry.trait)
        if ek in key or key in ek:
            return entry
    return None


def evidence_supported_by_review(evidence: str, review: str) -> bool:
    ev = evidence.strip().casefold()
    rv = review.strip().casefold()
    if not ev or not rv:
        return False
    if ev in rv:
        return True
    words = re.findall(r"[a-z0-9']{4,}", ev, flags=re.I)
    if not words:
        return False
    matched = [w for w in words if w.casefold() in rv]
    return len(matched) >= min(2, len(words))


def clamp_confidence(value: Any) -> float:
    try:
        num = float(value)
    except (TypeError, ValueError):
        return 0.65
    return max(0.0, min(1.0, num))


def short_trait_label(trait: str) -> str:
    trimmed = trait.strip()
    if len(trimmed) <= 72:
        return trimmed
    return trimmed[:69].rstrip() + "…"


def build_fallback_summary(influences: list[dict[str, Any]], _tribe_name: str) -> str:
    if not influences:
        return (
            "Tribe traits and user traits shaped what this review emphasizes — "
            "which product points get airtime, how strongly they're rated, and the overall tone."
        )
    tribe_hits = [i for i in influences if i.get("source") == "tribe"]
    user_hits = [i for i in influences if i.get("source") == "user"]
    top = influences[0]
    parts: list[str] = []
    if tribe_hits:
        parts.append(
            f"tribe traits around {short_trait_label(str(tribe_hits[0].get('trait') or '')).lower()} "
            "steer the review toward matching priorities"
        )
    if user_hits:
        parts.append(
            f"user traits around {short_trait_label(str(user_hits[0].get('trait') or '')).lower()} "
            "sharpen tone and emphasis"
        )
    joined = ", while ".join(parts) if parts else "known persona traits steer the review's focus"
    joined = joined[0].upper() + joined[1:] if joined else joined
    conf = int(round(float(top.get("confidence") or 0) * 100))
    return (
        f"{joined}. "
        f"The clearest influence ({conf}% confidence) is {short_trait_label(str(top.get('trait') or '')).lower()}, "
        "which shows up directly in what the review chooses to praise or criticize."
    )


def sanitize_influences(
    influences: list[dict[str, Any]],
    catalog: list[TraitCatalogEntry],
) -> list[dict[str, Any]]:
    seen: set[str] = set()
    cleaned: list[dict[str, Any]] = []
    for item in influences:
        trait = str(item.get("trait") or "").strip()
        evidence = str(item.get("evidence") or "").strip()
        confidence = clamp_confidence(item.get("confidence"))
        if not trait or not evidence or confidence < 0.6:
            continue
        match = resolve_catalog_entry(trait, catalog)
        if not match:
            continue
        dedupe = f"{match.source}:{normalize_trait_key(match.trait)}"
        if dedupe in seen:
            continue
        seen.add(dedupe)
        cleaned.append(
            {
                "trait": match.trait,
                "source": match.source,
                "traitGroup": match.trait_group,
                "evidence": evidence,
                "confidence": round(confidence, 2),
            }
        )
    cleaned.sort(key=lambda x: x.get("confidence", 0), reverse=True)
    return cleaned[:5]


def sanitize_summary(summary: str, influences: list[dict[str, Any]], tribe_name: str) -> str:
    cleaned = " ".join(str(summary or "").split()).strip()
    if not cleaned or len(cleaned) < 80:
        return build_fallback_summary(influences, tribe_name)
    return cleaned


def fallback_influences(
    sapiens_review: str,
    catalog: list[TraitCatalogEntry],
) -> list[dict[str, Any]]:
    review = sapiens_review.lower()
    hits: list[dict[str, Any]] = []
    for entry in catalog:
        words = re.findall(r"[a-z]{5,}", entry.trait.lower())[:8]
        matched = [w for w in words if w in review]
        if len(matched) < 2:
            continue
        hits.append(
            {
                "trait": entry.trait,
                "source": entry.source,
                "traitGroup": entry.trait_group,
                "evidence": (
                    f"The review's focus and wording reflect this trait — especially around "
                    f"{', '.join(matched[:3])}."
                ),
                "confidence": round(min(0.72, 0.45 + len(matched) * 0.08), 2),
            }
        )
        if len(hits) >= 3:
            break
    return hits


def parse_inference_json(raw: str) -> tuple[str, list[dict[str, Any]]] | None:
    try:
        start = raw.index("{")
        end = raw.rindex("}")
        parsed = json.loads(raw[start : end + 1])
        influences = parsed.get("influences")
        if not isinstance(influences, list):
            return None
        summary = str(parsed.get("summary") or "").strip()
        return summary, influences
    except (ValueError, json.JSONDecodeError):
        return None


def build_prompt(
    tribe_name: str,
    tribe_block: str,
    user_block: str,
    sapiens_review: str,
) -> tuple[str, str]:
    system = (
        "You identify which known tribe and user traits shaped a generated product review's opinion. "
        "Explain influence causally — what traits pulled the review toward certain priorities, tone, and emphasis. "
        "Use ONLY traits from the provided lists. Every inference must cite concrete evidence from the review. "
        "Assign honest confidence scores. Do not invent traits. Do not reference ground-truth reviews or user norms. Output only JSON."
    )
    prompt = f"""Tribe: {tribe_name}

TRIBE TRAITS (choose only from this list):
{tribe_block or "(none)"}

USER TRAITS (choose only from this list):
{user_block}

SAPIENS GENERATED REVIEW:
{sapiens_review}

Return JSON exactly:
{{
  "summary": "<2-3 sentence client-facing explanation>",
  "influences": [
    {{
      "trait": "<exact trait text from the lists above>",
      "source": "tribe" | "user",
      "traitGroup": "<tribe group label if tribe trait, otherwise User traits>",
      "evidence": "<short explanation of how this trait shows up in the review — what the review emphasizes and why that reflects the trait>",
      "confidence": <number 0.0-1.0>
    }}
  ]
}}

Summary rules (this is the main narrative the client reads):
- Write 2-3 sentences explaining HOW tribe traits and user traits influenced this review's opinion — not how SAPIENS works.
- Only mention traits that clearly helped shape the review. Omit any trait that did not materially influence the opinion.
- The summary must only discuss traits you include in the influences list — do not name extra traits.
- Describe causally: because tribe traits prioritize X, the review emphasizes Y; because user traits emphasize Z, the tone/focus shifts toward W.
- Always say "tribe traits" and "user traits" — never "the user", "this user", or "the reviewer" when referring to persona traits.
- Do NOT say "SAPIENS analyzes", "through the lens of", or repeat the tribe name as a label.
- Do NOT describe the inference process or methodology.
- Do NOT use bullet points.

Influence + confidence rules:
- Include only traits that clearly helped shape the review — quality over quantity (typically 2-4).
- Omit weak, tangential, or non-contributing traits entirely (do not list them).
- evidence should explain HOW the trait influenced the review — what the review emphasizes, praises, or criticizes, and why that reflects this trait. Ground it in the review but write analytically, not as a bare quote.
- In evidence, say "tribe traits" or "user traits" — not "the user".
- confidence guide:
  - 0.85-1.0 = trait clearly drives a major theme in the review
  - 0.65-0.84 = trait strongly visible in tone or focus
  - below 0.65 = do not include
- Prefer a mix of tribe traits + user traits when both clearly contributed.
- Skip traits with no support in the review."""
    return system, prompt


def call_openai(system: str, prompt: str, *, model: str, temperature: float) -> str:
    try:
        from openai import OpenAI
    except ImportError as exc:
        raise RuntimeError("Install openai: pip install openai") from exc

    api_key = os.environ.get("OPENAI_API_KEY", "").strip()
    if not api_key:
        raise RuntimeError("OPENAI_API_KEY is not set")

    client = OpenAI(api_key=api_key)
    response = client.chat.completions.create(
        model=model,
        messages=[
            {"role": "system", "content": system},
            {"role": "user", "content": prompt},
        ],
        temperature=temperature,
    )
    return (response.choices[0].message.content or "").strip()


def infer_trait_influences(
    row: CatalogProductRow,
    sub_to_main: dict[str, str],
    *,
    use_llm: bool,
    model: str,
) -> tuple[str, list[dict[str, Any]], str]:
    review = row.sapiens_review.strip()
    if not review:
        return "", [], "empty_review"

    traits = extract_qualitative_traits(row.qualitative_summary)
    user_block = format_user_characteristics(
        row.user_characteristic_summary,
        row.category_characteristics,
        row.category,
        sub_to_main,
    )
    tribe_block, user_block, catalog = build_trait_catalog(
        traits,
        user_block,
        row.user_characteristic_summary,
        row.category,
        row.category_characteristics,
        sub_to_main,
    )
    if not catalog:
        return "", [], "no_traits"

    if not use_llm:
        influences = fallback_influences(review, catalog)
        summary = build_fallback_summary(influences, row.tribe_name)
        return summary, influences, "fallback"

    system, prompt = build_prompt(row.tribe_name, tribe_block, user_block, review)
    try:
        raw = call_openai(system, prompt, model=model, temperature=0.2)
        parsed = parse_inference_json(raw)
        if not parsed:
            influences = fallback_influences(review, catalog)
            return build_fallback_summary(influences, row.tribe_name), influences, "fallback_parse"
        summary_raw, parsed_influences = parsed
        cleaned = sanitize_influences(parsed_influences, catalog)
        influences = cleaned if cleaned else fallback_influences(review, catalog)
        summary = sanitize_summary(summary_raw, influences, row.tribe_name)
        source = "llm" if cleaned else "fallback_sanitize"
        return summary, influences, source
    except Exception as exc:
        print(f"  LLM error for {row.review_key}: {exc}", file=sys.stderr)
        influences = fallback_influences(review, catalog)
        return build_fallback_summary(influences, row.tribe_name), influences, "fallback_error"


def load_blind_run_predictions() -> dict[str, str]:
    out: dict[str, str] = {}
    for path in VG_BLIND_RUN_PATHS.values():
        if not path.is_file():
            continue
        doc = load_json(path)
        for _uid, rows in (doc.get("user_predictions") or {}).items():
            if not isinstance(rows, list):
                continue
            for row in rows:
                if not isinstance(row, dict):
                    continue
                rk = str(row.get("review_key") or "").strip()
                pred = row.get("prediction") or {}
                text = str(pred.get("review_text") or "").strip()
                if rk and text:
                    out[rk] = text
    return out


def load_healthcare_predictions() -> dict[str, str]:
    if not HEALTHCARE_ACCURACY_PATH.is_file():
        return {}
    doc = load_json(HEALTHCARE_ACCURACY_PATH)
    out: dict[str, str] = {}
    for row in doc.get("reviews") or []:
        rk = str(row.get("review_key") or "").strip()
        pred = str((row.get("prediction") or {}).get("review") or "").strip()
        if rk and pred:
            out[rk] = pred
    return out


def load_catalog_rows(
    blind_preds: dict[str, str],
    hc_preds: dict[str, str],
    user_cat_chars: dict[str, dict[str, str]],
) -> list[CatalogProductRow]:
    catalog = load_json(DATA_DIR / "catalog-index.json")
    tribes_dir = DATA_DIR / "tribes"
    rows: list[CatalogProductRow] = []

    for tribe_meta in catalog.get("tribes") or []:
        tribe_id = tribe_meta["id"]
        tribe = load_json(tribes_dir / f"{tribe_id}.json")
        domain = tribe.get("domain", tribe_meta.get("domain", ""))
        cluster = tribe.get("cluster", tribe_meta.get("cluster", ""))
        micro_id = tribe.get("micro_id", tribe_meta.get("microId", ""))
        qualitative = tribe.get("qualitative_summary") or {}

        char_by_user = {
            m["user_id"]: m for m in tribe.get("member_user_characteristics") or []
        }

        for uid, products in (tribe.get("members_grouped_by_user") or {}).items():
            member = char_by_user.get(uid) or {}
            user_summary = str(member.get("characteristic_summary") or "").strip()
            cat_chars = member.get("category_characteristics") or {}
            if not cat_chars:
                cat_chars = domain_category_characteristics(uid, user_cat_chars, domain)

            for product in products:
                rk = str(product.get("review_key") or "").strip()
                if domain == "healthcare":
                    sapiens_review = hc_preds.get(rk) or str(product.get("user_history_review") or "").strip()
                else:
                    sapiens_review = blind_preds.get(rk) or str(product.get("user_history_review") or "").strip()

                rows.append(
                    CatalogProductRow(
                        domain=domain,
                        tribe_id=tribe_id,
                        tribe_name=tribe.get("tribe_name", tribe_meta.get("name", "")),
                        cluster=cluster,
                        micro_id=micro_id,
                        user_id=uid,
                        review_key=rk,
                        product_description=str(product.get("product_description") or "").strip(),
                        category=str(product.get("category") or "").strip(),
                        sapiens_review=sapiens_review,
                        user_characteristic_summary=user_summary,
                        category_characteristics=cat_chars if isinstance(cat_chars, dict) else {},
                        qualitative_summary=qualitative,
                        sapiens_baseline_gap=product.get("sapiens_baseline_gap"),
                        overall_similarity_score=product.get("overall_similarity_score"),
                    )
                )
    return rows


def sort_rows(rows: list[CatalogProductRow]) -> list[CatalogProductRow]:
    return sorted(
        rows,
        key=lambda r: (
            -(r.sapiens_baseline_gap or 0),
            -(r.overall_similarity_score or 0),
            r.domain,
            r.tribe_name,
            r.review_key,
        ),
    )


def write_json_output(rows: list[CatalogProductRow], path: Path) -> None:
    payload = {
        "generated_at": datetime.now().isoformat(timespec="seconds"),
        "model": INFERRED_TRAITS_MODEL,
        "product_count": len(rows),
        "products": [
            {
                "domain": r.domain,
                "tribeId": r.tribe_id,
                "tribeName": r.tribe_name,
                "cluster": r.cluster,
                "microId": r.micro_id,
                "userId": r.user_id,
                "reviewKey": r.review_key,
                "productDescription": r.product_description,
                "category": r.category,
                "sapiensReview": r.sapiens_review,
                "userCharacteristicSummary": r.user_characteristic_summary,
                "categoryCharacteristics": r.category_characteristics,
                "sapiensBaselineGap": r.sapiens_baseline_gap,
                "overallSimilarityScore": r.overall_similarity_score,
                "inferenceSource": r.inference_source,
                "summary": r.summary,
                "influences": r.influences,
            }
            for r in rows
        ],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def write_excel_output(rows: list[CatalogProductRow], path: Path) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Inferred Traits"

    headers = [
        "Domain",
        "Tribe Name",
        "User ID",
        "Review Key",
        "Product Description",
        "Category",
        "Sapiens Baseline Gap",
        "Overall Similarity",
        "Sapiens Review",
        "User Summary",
        "Category Characteristics",
        "Inference Source",
        "Summary",
        "Influence #",
        "Trait",
        "Source",
        "Trait Group",
        "Confidence",
        "Evidence",
    ]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    wrap = Alignment(wrap_text=True, vertical="top")
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for r in rows:
        cat_chars = "\n\n".join(f"[{k}] {v}" for k, v in r.category_characteristics.items())
        base = [
            r.domain,
            r.tribe_name,
            r.user_id,
            r.review_key,
            r.product_description,
            r.category,
            r.sapiens_baseline_gap,
            r.overall_similarity_score,
            r.sapiens_review,
            r.user_characteristic_summary,
            cat_chars,
            r.inference_source,
            r.summary,
        ]
        if not r.influences:
            ws.append(base + ["", "", "", "", "", ""])
            continue
        for idx, inf in enumerate(r.influences, start=1):
            ws.append(
                base
                + [
                    idx,
                    inf.get("trait", ""),
                    inf.get("source", ""),
                    inf.get("traitGroup", ""),
                    inf.get("confidence"),
                    inf.get("evidence", ""),
                ]
            )

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = wrap

    widths = {
        "A": 14,
        "B": 28,
        "E": 45,
        "I": 50,
        "J": 40,
        "K": 40,
        "M": 55,
        "O": 45,
        "P": 12,
        "S": 50,
    }
    for col, width in widths.items():
        ws.column_dimensions[col].width = width
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        if letter not in widths:
            ws.column_dimensions[letter].width = 18

    ws.freeze_panes = "A2"
    wb.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Batch infer trait influences for UI catalog products.")
    parser.add_argument(
        "--output-json",
        type=Path,
        default=ROOT / "inferred_trait_influences.json",
        help="JSON output path",
    )
    parser.add_argument(
        "--output-xlsx",
        type=Path,
        default=ROOT / "inferred_trait_influences.xlsx",
        help="Excel output path",
    )
    parser.add_argument("--limit", type=int, default=0, help="Process only first N products (0 = all)")
    parser.add_argument("--dry-run", action="store_true", help="Use keyword fallback only (no OpenAI calls)")
    parser.add_argument("--sleep", type=float, default=0.25, help="Seconds between API calls")
    parser.add_argument("--model", default=INFERRED_TRAITS_MODEL, help="OpenAI model name")
    args = parser.parse_args()

    sub_to_main = load_category_mapping()
    user_cat_chars = load_user_category_characteristics()
    blind_preds = load_blind_run_predictions()
    hc_preds = load_healthcare_predictions()

    rows = sort_rows(load_catalog_rows(blind_preds, hc_preds, user_cat_chars))
    if args.limit > 0:
        rows = rows[: args.limit]

    use_llm = not args.dry_run
    if use_llm and not os.environ.get("OPENAI_API_KEY", "").strip():
        print("OPENAI_API_KEY not set — falling back to keyword heuristic (--dry-run mode).", file=sys.stderr)
        use_llm = False

    print(f"Processing {len(rows)} catalog products ({'gpt-4o-mini' if use_llm else 'fallback'})…")

    for i, row in enumerate(rows, start=1):
        summary, influences, source = infer_trait_influences(
            row,
            sub_to_main,
            use_llm=use_llm,
            model=args.model,
        )
        row.summary = summary
        row.influences = influences
        row.inference_source = source
        print(
            f"[{i}/{len(rows)}] {row.tribe_name[:30]:30} | {row.review_key} | "
            f"{len(influences)} influences ({source})"
        )
        if use_llm and i < len(rows) and args.sleep > 0:
            time.sleep(args.sleep)

    write_json_output(rows, args.output_json)
    write_excel_output(rows, args.output_xlsx)
    print(f"\nWrote JSON: {args.output_json}")
    print(f"Wrote Excel: {args.output_xlsx}")


if __name__ == "__main__":
    main()
